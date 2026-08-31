# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Cotización"

HEAD_FILL = PatternFill("solid", fgColor="1F2933")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
CAT_FILL = PatternFill("solid", fgColor="EDE9E2")
CAT_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="D9D3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(size=10, italic=True, color="666666")

ws.merge_cells("A1:G1")
ws["A1"] = "Módulo sanitario — Proyecto El Túnel — Cotización de materiales"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:G2")
ws["A2"] = "9,83 x 5,00 m · dos baños de hombres espejados · minimalista industrial (microcemento / epóxico / concreto pulido)"
ws["A2"].font = SUB_FONT

headers = ["Categoría", "Material / producto", "Cantidad", "Unidad", "Precio unitario", "Precio total", "Nota"]
row0 = 4
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=row0, column=i, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

rows = [
    ("Piso", "Pintura epóxica gris cemento + imprimante", 48.40, "m²", "2 manos, antideslizante en zona húmeda"),
    ("Piso", "Rejilla de piso + sifón", 2, "un", "Una por módulo, pendiente 1,5–2 %"),
    ("Paredes", "Kit microcemento (base + microcemento + sellador)", 107.30, "m²", "3 caras existentes + tabique (2 caras) + frente nuevo"),
    ("Tabiquería", "Panel fenólico HPL 12–13 mm + herrajes inox.", 6, "cubículos", "Piso a techo, sin zócalo"),
    ("Tabiquería", "Puerta de cubículo fenólico 0,90 x 1,80 m", 6, "un", "Con pestillo y tope"),
    ("Mesón", "Concreto pulido + hidrofugante", 4.08, "m²", "2,64 m² tope + 1,44 m² frente descolgado"),
    ("Sanitarios", "Poceta (sanitario) loza blanca", 6, "un", "Cubículo 0,90 x 1,50 m"),
    ("Sanitarios", "Urinario suspendido loza blanca", 8, "un", "Separación 0,75 m a eje"),
    ("Sanitarios", "Lavamanos empotrado loza blanca", 6, "un", "Separación 0,80 m a eje"),
    ("Grifería", "Llave de lavamanos, negro mate", 6, "un", ""),
    ("Grifería", "Fluxómetro / válvula de poceta", 6, "un", ""),
    ("Grifería", "Válvula de urinario", 8, "un", ""),
    ("Espejos", "Espejo corrido 4 mm, 2,40 m", 2, "un", "Uno por mesón"),
    ("Mobiliario", "Banco 1,60 x 0,45 m, concreto pulido o metal + madera", 2, "un", "Esquina liberada de cada módulo"),
    ("Puertas de acceso", "Puerta con cerradura, acabado negro mate", 2, "un", "Una por módulo"),
    ("Accesorios", "Dispensador de papel", 2, "un", "1 por módulo, a confirmar cantidad"),
    ("Accesorios", "Dispensador de jabón", 2, "un", "1 por módulo, a confirmar cantidad"),
    ("Accesorios", "Basurero", 2, "un", "1 por módulo, a confirmar cantidad"),
    ("Obra civil", "Bloque + friso, pared frontal nueva", 1, "global", "Única pared nueva; las otras 3 ya existen"),
    ("Instalaciones", "Acometida de aguas blancas + descarga a cloaca existente", 1, "global", "Por el frente del módulo, según levantamiento"),
]

r = row0 + 1
for cat, mat, qty, unit, note in rows:
    ws.cell(row=r, column=1, value=cat).font = CAT_FONT
    ws.cell(row=r, column=1).fill = CAT_FILL
    ws.cell(row=r, column=2, value=mat)
    ws.cell(row=r, column=3, value=qty).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=4, value=unit).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5, value=None)  # precio unitario a llenar
    ws.cell(row=r, column=6, value=f"=IF(E{r}=\"\",\"\",C{r}*E{r})")
    ws.cell(row=r, column=7, value=note)
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = BORDER
    r += 1

total_row = r + 1
ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
ws.cell(row=total_row, column=6, value=f"=SUM(F{row0+1}:F{r-1})").font = Font(bold=True)
for col in (5, 6):
    ws.cell(row=total_row, column=col).border = BORDER

widths = [16, 46, 10, 10, 14, 14, 42]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# ---- segunda hoja: dimensiones y cómputo verificado, para referencia ----
ws2 = wb.create_sheet("Dimensiones verificadas")
ws2.append(["Elemento", "Valor"])
for c in ws2[1]:
    c.font = HEAD_FONT; c.fill = HEAD_FILL
datos = [
    ("Ancho libre entre muros", "9,83 m"),
    ("Fondo", "5,00 m"),
    ("Área bruta", "49,15 m²"),
    ("Tabique central", "0,15 m"),
    ("Ancho por módulo", "4,84 m"),
    ("Área por módulo", "24,20 m²"),
    ("Pocetas por módulo", "3 (0,90 x 1,50 m c/u, pegadas al tabique)"),
    ("Urinarios por módulo", "4 (0,75 m a eje, 3,00 m corrido, paño del tabique)"),
    ("Lavamanos por módulo", "3 (mesón corrido 2,40 m, muro exterior)"),
    ("Banco por módulo", "1,60 x 0,45 m, esquina liberada junto al muro exterior"),
    ("Puertas de acceso", "extremos opuestos del frente (izquierda y derecha)"),
    ("Total piezas sanitarias", "20 (6 pocetas + 8 urinarios + 6 lavamanos)"),
    ("Método de verificación", "shapely (área por polígono + comprobación de solapes)"),
]
for k, v in datos:
    ws2.append([k, v])
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 55

wb.save("/tmp/claude-0/-home-user-produccion-studio/344e1d7f-bb05-50a6-8219-29e2715272fc/scratchpad/El_Tunel_cotizacion_materiales.xlsx")
print("xlsx guardado")
